"""
CPS Invoice Upload Template Automation
======================================
Reads CPS_Database.xlsx (Fixed + App Inputs sheets) and populates
the Upload_Template.xlsx with the correct billing data for the current month.

Logic:
- Match Contract Name across Fixed and App Inputs sheets (fuzzy match handles
  minor spelling differences like "West Minister" vs "Westminster")
- Check frequency + billing type to determine which month column to read
- Only process rows where current-month net value is present (non-NaN, non-zero)
- Build text lines dynamically based on frequency and billing type
- Populate Upload Template with one H row + N L rows per contract

Run:
    python cps_invoice_automation.py
    python cps_invoice_automation.py /path/to/folder   # if files are elsewhere
"""

import os
import sys
import shutil
import re
import datetime
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from dateutil.relativedelta import relativedelta

# ─────────────────────────────────────────────────────────────────────────────
# CONFIG  — change only these paths if needed
# ─────────────────────────────────────────────────────────────────────────────
CPS_DB_PATH   = "CPS_Database__1_.xlsx"
TEMPLATE_PATH = "Upload_Template.xlsx"
OUTPUT_PATH   = "Populated_Upload_Template.xlsx"

FIXED_SHEET   = "Fixed"
APP_SHEET     = "App Inputs"

# ─────────────────────────────────────────────────────────────────────────────
# TEMPLATE COLUMN ORDER  (must match Upload Template row 1 exactly)
# ─────────────────────────────────────────────────────────────────────────────
TEMPLATE_HEADERS = [
    "Line Type Indicator", "Line Type Key", "Order Type", "Bill-To", "Sold-To",
    "Customer Name", "PO Number", "Contact Name", "Invoice Date", "Payment Days",
    "Direct Debit", "Currency", "Order_Reason", "Capita Contact Name",
    "Capita Contact Number", "Material Code", "Net Value", "VAT Indicator",
    "Release From", "Release Until", "WBS",
    "Text Line 1", "Text Line 2", "Text Line 3", "Text Line 4", "Text Line 5",
]

# ─────────────────────────────────────────────────────────────────────────────
# DATE UTILITIES
# ─────────────────────────────────────────────────────────────────────────────

def get_current_month() -> datetime.date:
    """Return first day of the current month. Override here for testing."""
    return datetime.date.today().replace(day=1)


def col_label(dt: datetime.date) -> str:
    """date → column header like 'May-26'."""
    return dt.strftime("%b-%y")


def friendly(dt: datetime.date) -> str:
    """date → display label like 'May 26'."""
    return dt.strftime("%b %y")


def parse_col(label: str) -> datetime.date:
    """'May-26' → datetime.date(2026, 5, 1)."""
    return datetime.datetime.strptime(label, "%b-%y").date().replace(day=1)


QUARTER_STARTS = (1, 4, 7, 10)


def quarter_start_for(dt: datetime.date) -> datetime.date:
    """Return the first month of the quarter containing dt."""
    for qs in reversed(QUARTER_STARTS):
        if dt.month >= qs:
            return dt.replace(month=qs, day=1)
    return dt.replace(month=1, day=1)


def next_quarter_start(dt: datetime.date) -> datetime.date:
    """Return the first month of the NEXT quarter after dt."""
    for qs in QUARTER_STARTS:
        if dt.month < qs:
            return dt.replace(month=qs, day=1)
    return dt.replace(year=dt.year + 1, month=1, day=1)


def prev_quarter_start(dt: datetime.date) -> datetime.date:
    """Return the first month of the quarter BEFORE the one containing dt."""
    cur_qs = quarter_start_for(dt)
    return (cur_qs - relativedelta(months=3)).replace(day=1)


def quarter_end(q_start: datetime.date) -> datetime.date:
    """Return the third month of a quarter."""
    return q_start + relativedelta(months=2)


# ─────────────────────────────────────────────────────────────────────────────
# BILLING PERIOD LABEL
# ─────────────────────────────────────────────────────────────────────────────

def billing_period_label(frequency: str, billing_type: str,
                         data_month: datetime.date) -> str:
    """
    Build the period text to append to the first text line.

    frequency    : Monthly | Quarterly | Annually  (case-insensitive)
    billing_type : In-month | Advance | Arrear/Arrears  (case-insensitive)
    data_month   : the month column from which the net value was taken
    """
    freq  = frequency.strip().lower()
    btype = billing_type.strip().lower().rstrip("s")   # normalise 'arrears' → 'arrear'

    if freq == "monthly":
        if btype == "in-month":
            label_dt = data_month
        elif btype == "advance":
            label_dt = data_month + relativedelta(months=1)
        elif btype == "arrear":
            label_dt = data_month - relativedelta(months=1)
        else:
            label_dt = data_month
        return friendly(label_dt)

    elif freq == "quarterly":
        if btype == "in-month":
            q_start = quarter_start_for(data_month)
        elif btype == "advance":
            q_start = next_quarter_start(data_month)
        elif btype == "arrear":
            q_start = prev_quarter_start(data_month)
        else:
            q_start = quarter_start_for(data_month)
        return f"{friendly(q_start)} to {friendly(quarter_end(q_start))}"

    else:
        # Annually — no period label for now (per requirements)
        return ""


# ─────────────────────────────────────────────────────────────────────────────
# DATA LOADING
# ─────────────────────────────────────────────────────────────────────────────

def load_fixed(path: str) -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name=FIXED_SHEET, header=0)
    df.columns = [str(c).strip() for c in df.columns]
    df["Contract Name"] = df["Contract Name"].astype(str).str.strip()
    return df


def load_app(path: str):
    """Returns (DataFrame, [month_col_labels])."""
    df = pd.read_excel(path, sheet_name=APP_SHEET, header=0)
    df.columns = [str(c).strip() for c in df.columns]
    df.rename(columns={df.columns[0]: "Contract name"}, inplace=True)
    df["Contract name"] = df["Contract name"].astype(str).str.strip()

    month_re = re.compile(r'^[A-Z][a-z]{2}-\d{2}$')
    month_cols = [c for c in df.columns if month_re.match(c)]
    return df, month_cols


# ─────────────────────────────────────────────────────────────────────────────
# FUZZY NAME MATCHING
# ─────────────────────────────────────────────────────────────────────────────

def normalise(name: str) -> str:
    """Lower-case, remove spaces and hyphens for fuzzy comparison."""
    return re.sub(r'[\s\-]', '', name.lower())


def build_fixed_lookup(fixed_df: pd.DataFrame) -> dict:
    """
    Returns dict: normalised_name → pd.Series (the fixed row).
    Handles 'West Minister' → 'westminster' etc.
    """
    lookup = {}
    for _, row in fixed_df.iterrows():
        key = normalise(row["Contract Name"])
        lookup[key] = row
    return lookup


def match_fixed(contract_name: str, lookup: dict):
    """
    Try exact normalised match, substring match, then similarity match.
    Returns the fixed-sheet Series or None.
    """
    from difflib import SequenceMatcher
    key = normalise(contract_name)

    # 1. Exact normalised match
    if key in lookup:
        return lookup[key]

    # 2. Substring match (one name contains the other)
    for fixed_key, row in lookup.items():
        if key in fixed_key or fixed_key in key:
            print(f"     ↳ substring match: '{contract_name}' ↔ '{row['Contract Name']}'")
            return row

    # 3. High-similarity match (handles typos like Westminster vs West Minister)
    SIMILARITY_THRESHOLD = 0.85
    best_score, best_row = 0, None
    for fixed_key, row in lookup.items():
        score = SequenceMatcher(None, key, fixed_key).ratio()
        if score > best_score:
            best_score, best_row = score, row
    if best_score >= SIMILARITY_THRESHOLD:
        print(f"     ↳ similarity match ({best_score:.2f}): '{contract_name}' ↔ '{best_row['Contract Name']}'")
        return best_row

    return None


# ─────────────────────────────────────────────────────────────────────────────
# ACTIVE MONTH DETECTION
# ─────────────────────────────────────────────────────────────────────────────

def active_col_for(app_group: pd.DataFrame, month_cols: list,
                   current_month: datetime.date):
    """
    Return the column label for current_month if it has at least one
    non-NaN, non-zero value in the group. Otherwise return None.
    """
    target = col_label(current_month)
    if target in month_cols:
        vals = app_group[target]
        if vals.notna().any() and (vals.fillna(0) != 0).any():
            return target
    return None


# ─────────────────────────────────────────────────────────────────────────────
# PROCESS ONE CONTRACT
# ─────────────────────────────────────────────────────────────────────────────

def process_contract(contract_name: str,
                     app_group: pd.DataFrame,
                     fixed_row: pd.Series,
                     active_col: str) -> list:
    """
    Build L-row dicts for every App Inputs row in the group.
    """
    data_month   = parse_col(active_col)
    frequency    = str(app_group["Frequency"].iloc[0]).strip()
    billing_type = str(app_group["Billing type"].iloc[0]).strip()
    period_lbl   = billing_period_label(frequency, billing_type, data_month)

    def _clean(v):
        return None if (v is None or (isinstance(v, float) and pd.isna(v))) else v

    rows_out = []
    for pos, (_, app_row) in enumerate(app_group.iterrows()):
        net_val  = app_row.get(active_col)
        text_raw = str(app_row.get("Lext lines", "")).strip()
        if text_raw in ("nan", "None", ""):
            text_raw = ""

        row = {
            "Line Type Indicator": "L",
            "Line Type Key":       1,
            "Order Type":          None,
            "Bill-To":             _clean(fixed_row.get("SAP ID")),
            "Sold-To":             _clean(fixed_row.get("SAP ID")),
            "Customer Name":       _clean(fixed_row.get("Customer Name")),
            "PO Number":           _clean(app_row.get("PO Number")) or _clean(fixed_row.get("PO Number")),
            "Contact Name":        _clean(fixed_row.get("Contact Name")),
            "Invoice Date":        None,
            "Payment Days":        _clean(fixed_row.get("Payment Days")),
            "Direct Debit":        None,
            "Currency":            _clean(fixed_row.get("Currency")),
            "Order_Reason":        None,
            "Capita Contact Name": _clean(fixed_row.get("Capita Contact Name")),
            "Capita Contact Number": _clean(fixed_row.get("Capita Contact Number")),
            "Material Code":       _clean(fixed_row.get("Material Code")),
            "Net Value":           _clean(net_val) if pd.notna(net_val) else 0,
            "VAT Indicator":       _clean(fixed_row.get("VAT Indicator")),
            "Release From":        _clean(fixed_row.get("Release From")),
            "Release Until":       _clean(fixed_row.get("Release Until")),
            "WBS":                 _clean(fixed_row.get("WBS")),
            "Text Line 1":         text_raw,
            "Text Line 2":         None,
            "Text Line 3":         None,
            "Text Line 4":         None,
            "Text Line 5":         None,
        }
        rows_out.append(row)

    # Append period label to FIRST row's Text Line 1 only
    if rows_out and period_lbl:
        t = rows_out[0]["Text Line 1"]
        rows_out[0]["Text Line 1"] = (f"{t} - {period_lbl}" if t else period_lbl).strip(" -")

    return rows_out


# ─────────────────────────────────────────────────────────────────────────────
# WRITE OUTPUT
# ─────────────────────────────────────────────────────────────────────────────

def write_output(all_results: list, template_path: str, output_path: str):
    """
    Copy template, clear data rows (keep header row 1),
    then write H + L rows for each processed contract.
    """
    shutil.copy(template_path, output_path)
    wb = load_workbook(output_path)
    ws = wb.active

    # Clear data rows (keep header in row 1)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.value = None

    write_row = 2

    def _put(row_num, data):
        for col_idx, header in enumerate(TEMPLATE_HEADERS, start=1):
            val = data.get(header)
            ws.cell(row=row_num, column=col_idx).value = (
                None if (val is None or (isinstance(val, float) and pd.isna(val))) else val
            )

    for contract_name, l_rows in all_results:
        if not l_rows:
            continue
        first = l_rows[0]

        # ── H Row ──────────────────────────────────────────────────────────
        h = {
            "Line Type Indicator": "H",
            "Line Type Key":       1,
            "Order Type":          first.get("Order Type"),
            "Bill-To":             first.get("Bill-To"),
            "Sold-To":             first.get("Sold-To"),
            "Customer Name":       first.get("Customer Name"),
            "PO Number":           first.get("PO Number"),
            "Contact Name":        first.get("Contact Name"),
            "Invoice Date":        first.get("Invoice Date"),
            "Payment Days":        first.get("Payment Days"),
            "Direct Debit":        first.get("Direct Debit"),
            "Currency":            first.get("Currency"),
            "Order_Reason":        first.get("Order_Reason"),
            "Capita Contact Name": first.get("Capita Contact Name"),
            "Capita Contact Number": first.get("Capita Contact Number"),
            "Material Code":       None,
            "Net Value":           None,
            "VAT Indicator":       None,
            "Release From":        None,
            "Release Until":       None,
            "WBS":                 None,
            "Text Line 1":         contract_name,
            "Text Line 2": None, "Text Line 3": None,
            "Text Line 4": None, "Text Line 5": None,
        }
        _put(write_row, h)
        write_row += 1

        for r in l_rows:
            _put(write_row, r)
            write_row += 1

    wb.save(output_path)
    total_data = write_row - 2
    print(f"\n{'='*60}")
    print(f"✅  Output saved → {output_path}")
    print(f"   Data rows written: {total_data}  (excl. header)")
    print(f"{'='*60}")


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("CPS Invoice Upload Template Automation")
    print("=" * 60)

    current_month = get_current_month()
    print(f"Current month   : {col_label(current_month)}  ({current_month})")

    fixed_df          = load_fixed(CPS_DB_PATH)
    app_df, month_cols = load_app(CPS_DB_PATH)
    fixed_lookup      = build_fixed_lookup(fixed_df)

    print(f"Fixed contracts : {len(fixed_df)}")
    print(f"App Input rows  : {len(app_df)}  |  Month columns: {month_cols}\n")

    # Preserve original order of first appearance per contract
    seen = {}
    for name in app_df["Contract name"]:
        if name not in seen and name.lower() not in ("nan", "none", ""):
            seen[name] = True
    contract_order = list(seen.keys())

    all_results = []
    skipped     = []

    for contract_name in contract_order:
        app_group = app_df[app_df["Contract name"] == contract_name].copy()

        # Gate: must have current-month data
        a_col = active_col_for(app_group, month_cols, current_month)
        if a_col is None:
            print(f"  ⚠  SKIP '{contract_name}'  — no data in {col_label(current_month)}")
            skipped.append(contract_name)
            continue

        # Match to Fixed sheet
        fixed_row = match_fixed(contract_name, fixed_lookup)
        if fixed_row is None:
            print(f"  ⚠  SKIP '{contract_name}'  — not found in Fixed sheet")
            skipped.append(contract_name)
            continue

        rows = process_contract(contract_name, app_group, fixed_row, a_col)
        all_results.append((contract_name, rows))

        freq  = app_group["Frequency"].iloc[0]
        btype = app_group["Billing type"].iloc[0]
        period = billing_period_label(str(freq), str(btype), parse_col(a_col))
        print(f"  ✓  '{contract_name}'"
              f"  |  {freq} / {btype}"
              f"  |  col={a_col}"
              f"  |  period='{period}'"
              f"  |  {len(rows)} L-rows")

    print(f"\nProcessed : {len(all_results)} contracts")
    print(f"Skipped   : {len(skipped)} → {skipped}")

    write_output(all_results, TEMPLATE_PATH, OUTPUT_PATH)


if __name__ == "__main__":
    if len(sys.argv) > 1:
        os.chdir(sys.argv[1])
    main()
